"""
Excel report generator for Banner Billings.

This module generates formatted Excel workbooks from combined billing data,
including a summary sheet and a data sheet.
"""
import logging
from typing import Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from io import BytesIO
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Formatting constants
HEADER_BG_COLOR = "D9D9D9"
ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


class ExcelReportGenerator:
    """
    Generates formatted Excel reports from combined Banner Billings data.
    
    Creates a workbook with:
    1. Summary sheet - Processing metadata and statistics
    2. Data sheet - All combined billing rows
    
    Future enhancements:
    - Additional formatting (column widths, headers, number formats)
    - Multiple data sheets by category
    - Charts and visualizations
    """
    
    def __init__(self):
        """Initialize the report generator."""
        self.default_sheet_name = "Combined Data"
        self.summary_sheet_name = "Summary"
        self.study_summary_sheet_name = "Study Summary"
        self.account_summary_sheet_name = "Account Summary"

    def generate_report(
        self,
        data_df: pd.DataFrame,
        summary_data: Dict[str, Any],
        output_path: Optional[str] = None,
        study_summary_df: Optional[pd.DataFrame] = None,
        account_summary_df: Optional[pd.DataFrame] = None,
    ) -> BytesIO:
        """
        Generate an Excel report from the combined data.

        Args:
            data_df: DataFrame containing all combined billing rows
            summary_data: Dictionary containing summary statistics
            output_path: Optional path to save the file (also returns BytesIO)
            study_summary_df: Optional study-level summary (by Study Name, Study Code, etc.)
            account_summary_df: Optional account-level summary (by KFS No, Invoice Type, etc.)

        Returns:
            BytesIO object containing the Excel file
        """
        output_buffer = BytesIO()

        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            # Write Summary sheet first
            self._write_summary_sheet(writer, summary_data)

            # Write aggregated summary sheets (before Combined Data)
            if study_summary_df is not None and not study_summary_df.empty:
                study_order = [
                    'Study Name', 'Study Code', 'KFS No', 'IRB No',
                    'Invoice Date', 'Invoice Type',
                    'Charge Amount', 'Adjustment', 'Balance Due'
                ]
                study_df = study_summary_df[[c for c in study_order if c in study_summary_df.columns]]
                self._write_summary_table(writer, study_df, self.study_summary_sheet_name)
            if account_summary_df is not None and not account_summary_df.empty:
                # Reorder: Invoice Date before Invoice Type
                acct_order = [
                    'KFS No', 'Invoice Date', 'Invoice Type',
                    'Charge Amount', 'Adjustment', 'Balance Due'
                ]
                acct_df = account_summary_df[[c for c in acct_order if c in account_summary_df.columns]]
                self._write_summary_table(writer, acct_df, self.account_summary_sheet_name)

            # Write Data sheet
            self._write_data_sheet(writer, data_df)

            # Apply formatting to all sheets
            self._apply_formatting(
                writer,
                summary_data,
                study_summary_df,
                account_summary_df,
                data_df,
            )

        # Reset buffer position for reading
        output_buffer.seek(0)
        
        # Optionally save to file
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(output_buffer.getvalue())
            logger.info(f"Saved report to: {output_path}")
            output_buffer.seek(0)  # Reset again after writing
        
        logger.info(
            f"Generated Excel report: {len(data_df)} data rows, "
            f"{summary_data.get('total_sheets_processed', 0)} sheets processed"
        )
        
        return output_buffer
    
    def _write_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        summary_data: Dict[str, Any]
    ):
        """
        Write the summary sheet to the workbook.
        
        Args:
            writer: Excel writer object
            summary_data: Dictionary containing summary statistics
        """
        # Build summary rows
        summary_rows = [
            ['Banner Billings Report Summary', ''],
            ['', ''],
            ['Processing Information', ''],
            ['Generated', summary_data.get('processing_timestamp', 'N/A')],
            ['Total Source Files', summary_data.get('total_source_files', 0)],
            ['Total Sheets Processed', summary_data.get('total_sheets_processed', 0)],
            ['Total Rows Extracted', summary_data.get('total_rows_extracted', 0)],
            ['', ''],
            ['Source Files', ''],
        ]
        
        # Add source file list
        for filename in summary_data.get('source_files', []):
            summary_rows.append(['', filename])
        
        summary_rows.append(['', ''])
        summary_rows.append(['Sheet Details', ''])
        summary_rows.append([
            'Workbook', 'Sheet', 'Invoice Type', 'Invoice Date', 'PI', 'Study Name',
            'Study Code', 'IRB No', 'KFS No', 'Raw Rows', 'Extracted Rows'
        ])
        
        # Add per-sheet details (with extracted metadata)
        for sheet in summary_data.get('sheets', []):
            summary_rows.append([
                sheet.get('workbook', 'unknown'),
                sheet.get('sheet', 'unknown'),
                sheet.get('invoice_type', 'unknown'),
                sheet.get('invoice_date', ''),
                sheet.get('PI', ''),
                sheet.get('STUDY NAME', ''),
                sheet.get('STUDY CODE', ''),
                sheet.get('IRB NO', ''),
                sheet.get('KFS NO', ''),
                sheet.get('raw_rows', 0),
                sheet.get('extracted_rows', 0)
            ])
        
        # Convert to DataFrame and write
        # Use max columns across all rows
        max_cols = max(len(row) for row in summary_rows)
        padded_rows = [row + [''] * (max_cols - len(row)) for row in summary_rows]
        
        summary_df = pd.DataFrame(padded_rows)
        summary_df.to_excel(
            writer,
            sheet_name=self.summary_sheet_name,
            index=False,
            header=False
        )
        
        logger.debug("Wrote summary sheet")

    def _write_summary_table(
        self,
        writer: pd.ExcelWriter,
        df: pd.DataFrame,
        sheet_name: str,
    ) -> None:
        """Write a summary DataFrame to a sheet."""
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        logger.debug(f"Wrote {sheet_name}: {len(df)} rows")

    def _write_data_sheet(
        self,
        writer: pd.ExcelWriter,
        data_df: pd.DataFrame
    ):
        """
        Write the data sheet to the workbook.
        
        Args:
            writer: Excel writer object
            data_df: DataFrame containing all combined billing rows
        """
        if data_df.empty:
            # Write an empty sheet with a message
            empty_df = pd.DataFrame([['No data extracted']], columns=['Message'])
            empty_df.to_excel(
                writer,
                sheet_name=self.default_sheet_name,
                index=False
            )
            logger.warning("Data sheet is empty - no data extracted")
            return

        # Internal column names -> display names (title case, IRB/PI/KFS stay capitalized)
        rename_map = {
            '_workbook_name': 'Source Workbook',
            '_sheet_name': 'Source Sheet',
            '_invoice_type': 'Invoice Type',
            'invoice_date': 'Invoice Date',
            'STUDY NAME': 'Study Name',
            'PI': 'PI',
            'STUDY CODE': 'Study Code',
            'IRB NO': 'IRB No',
            'KFS NO': 'KFS No',
            'Charge Amount': 'Charge Amount',
            'Adjustment': 'Adjustment',
            'Balance Due': 'Balance Due',
        }
        # Desired column order on Combined Data tab (Invoice Date before Invoice Type)
        display_order = [
            'Source Workbook', 'Source Sheet', 'Invoice Date', 'Invoice Type',
            'Study Name', 'Study Code', 'PI', 'IRB No', 'KFS No',
            'Charge Amount', 'Adjustment', 'Balance Due',
        ]

        data_df = data_df.rename(columns=rename_map)
        # Reorder: only include columns that exist (in case some are missing)
        ordered_cols = [c for c in display_order if c in data_df.columns]
        # Append any extra columns not in display_order (e.g. future additions)
        extra_cols = [c for c in data_df.columns if c not in ordered_cols]
        data_df = data_df[ordered_cols + extra_cols]
        
        # Write to Excel
        data_df.to_excel(
            writer,
            sheet_name=self.default_sheet_name,
            index=False
        )
        
        logger.debug(f"Wrote data sheet: {len(data_df)} rows, {len(data_df.columns)} columns")

    def _apply_formatting(
        self,
        writer: pd.ExcelWriter,
        summary_data: Dict[str, Any],
        study_summary_df: Optional[pd.DataFrame],
        account_summary_df: Optional[pd.DataFrame],
        data_df: pd.DataFrame,
    ) -> None:
        """Apply formatting to all sheets in the workbook."""
        wb = writer.book
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True)
        right_alignment = Alignment(horizontal="right", wrap_text=True)

        # Summary sheet
        ws = wb[self.summary_sheet_name]
        num_source_files = len(summary_data.get("source_files", []))
        # Column names row: 9 (fixed) + num_source_files + 2 (blank + "Sheet Details") + 1 = 12 + num_source_files
        header_row = 12 + num_source_files
        num_sheet_rows = len(summary_data.get("sheets", []))
        last_table_row = header_row + num_sheet_rows

        # Bold labels in column A above the table (rows 1 to header_row-1)
        for row in range(1, header_row):
            cell = ws.cell(row=row, column=1)
            if cell.value:
                cell.font = bold_font

        # Bold + background + filters on table header row
        for col in range(1, 12):  # A through K
            cell = ws.cell(row=header_row, column=col)
            cell.font = bold_font
            cell.fill = header_fill

        ws.auto_filter.ref = f"A{header_row}:K{last_table_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        # Summary column widths
        summary_widths = [
            53.71, 23, 18.71, 14.57, 17.14, 22.43, 15.71, 14.71, 10.43, 9.71, 9.71
        ]
        for col, width in enumerate(summary_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Study Summary, Account Summary, Combined Data
        study_widths = {
            "Study Name": 22.43, "Study Code": 15.71, "KFS No": 10.43,
            "IRB No": 14.71, "Invoice Date": 14.57, "Invoice Type": 18.71,
            "Charge Amount": 16, "Adjustment": 16, "Balance Due": 16,
        }
        account_widths = {
            "KFS No": 10.43, "Invoice Date": 14.57, "Invoice Type": 18.71,
            "Charge Amount": 16, "Adjustment": 16, "Balance Due": 16,
        }
        combined_widths = {
            "Source Workbook": 53.71, "Source Sheet": 23, "Invoice Date": 14.57,
            "Invoice Type": 18.71, "Study Name": 22.43, "Study Code": 15.71,
            "PI": 18, "IRB No": 14.71, "KFS No": 10.43,
            "Charge Amount": 16, "Adjustment": 16, "Balance Due": 16,
        }

        for sheet_name, col_widths in [
            (self.study_summary_sheet_name, study_widths),
            (self.account_summary_sheet_name, account_widths),
            (self.default_sheet_name, combined_widths),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            self._format_data_sheet(
                ws, col_widths, bold_font, header_fill, wrap_alignment,
                right_alignment
            )

        logger.debug("Applied formatting to all sheets")

    def _format_data_sheet(
        self,
        ws,
        col_widths: Dict[str, float],
        bold_font: Font,
        header_fill: PatternFill,
        wrap_alignment: Alignment,
        right_alignment: Alignment,
    ) -> None:
        """Format a data sheet: bold header, freeze, filter, widths, wrap, accounting."""
        if ws.max_row < 1:
            return

        # Header row (row 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # Column widths
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val and str(header_val) in col_widths:
                ws.column_dimensions[col_letter].width = col_widths[str(header_val)]
            else:
                ws.column_dimensions[col_letter].width = 12

        # Data rows: wrap text, accounting format for Charge Amount/Adjustment/Balance Due
        amount_cols = {"Charge Amount", "Adjustment", "Balance Due"}
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = wrap_alignment
                header_val = ws.cell(row=1, column=col).value
                if header_val and str(header_val) in amount_cols:
                    cell.alignment = Alignment(horizontal="right", wrap_text=True)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = ACCOUNTING_FORMAT

    def generate_filename(self, prefix: str = "banner_billings") -> str:
        """
        Generate a timestamped filename for the report.
        
        Args:
            prefix: Filename prefix
            
        Returns:
            Filename string with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_report_{timestamp}.xlsx"
